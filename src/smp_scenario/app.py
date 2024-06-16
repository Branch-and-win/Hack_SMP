import os
import io
import shutil
import zipfile
from datetime import datetime

import openpyxl
import pandas as pd
from fastapi import FastAPI, UploadFile
from fastapi.responses import StreamingResponse

from src.smp_model.graph.length_velocity_calc import dump_velocity_length
from src.smp_scenario.parent_scenario import ParentScenario
from src.smp_scenario.parent_scenario_config import ParentScenarioConfig
from src.smp_scenario.scenario import Scenario
from src.smp_scenario.scenario_config import ScenarioConfig


description = """
API для сервиса по планированию маршрутов атомных ледоколов по СМП.

## Сценарии

**create_scenario** - 
Возможность создать сценарий для планирования маршрутов. 

При создании сценария есть возможность задать различные входные данные системы. Например, можно добавить или удалить
 заявки, скорректировать характеристики ледоколов и судов, изменить положение портов, добавить новые точки сбора
 караванов, изменить возможности перемещения между ребрами графа СМП. В запуск также передается информация о
 состоянии льда на каждом допустимом ребре СМП (информация об этом может быть сгенерирована с помощью другого вызова
 api системы calculate_ice_integrality.
 
Для сценария также можно выставить различные параметры моделирования. Можно настроить дату начала построения расписания,
 количество дней планирования, количество дней сверх плана, ограничение по времени на решение задачи и количество
 лучших маршрутов для построения расписания.

**create_parent_scenario** - 
Возможность создать родительский сценарий. Запуск такого сценария осуществляется последовательно интервалами времени,
 заданными в настройках сценария.

Для сценария можно указать последнюю дату планирования - дату до которой будет осуществлен последовательный расчет
 расписания.
 
**download_scenario**
Возможность выгрузить входные и выходные (если они есть) данные сценария.

## Моделирование

**calculate_scenario** - 
Возможность запустить выбранный сценарий для оптимизации расписания движения ледоколов по СМП. 

**calculate_ice_integrality** - 
Возможность запустить расчет тяжести льда.

При запуске необходимо передать данные о ледовой обстановке, данные о расположении портов, а также граф переходов в СМП.

## Отчет

После расчета сценария можно сформировать дашборд на основе выходных данных сценария. За формирования дашборда отвечает
 отдельный сервис dash_app, в котором реализован функционал загрузки результатов сценария для отображения отчета.
"""

app = FastAPI(
    title="🚢 BnW_SMP",
    description=description,
    summary="Приложение планирования маршрутов ледоколов в СМП команды Branch and Win.",
    version="0.0.1",
    contact={
        "name": "Branch and Win.",
        # "url": "http://x-force.example.com/contact/",
        "email": "ivs.rodin@gmail.com",
    },
    swagger_ui_parameters={"syntaxHighlight.theme": "obsidian"},
)


@app.post('/create_scenario/', tags=['Сценарии'])
async def create_scenario(
        name: str,
        main_data: UploadFile,
        velocity_env: UploadFile,
        start_date: str = '27-02-2022',
        duration_days: int = 7,
        # interval_hours: int = 1,
        cross_days: int = 2,
        timelimit: int = 900,
        k_bests: int = 5,
):
    """
    Создание сценария для планирования маршрутов.

    При создании сценария есть возможность задать различные входные данные системы. Например, можно добавить или удалить
     заявки, скорректировать характеристики ледоколов и судов, изменить положение портов, добавить новые точки сбора
     караванов, изменить возможности перемещения между ребрами графа СМП. В запуск также передается информация о
     состоянии льда на каждом допустимом ребре СМП (информация об этом может быть сгенерирована с помощью другого вызова
     api системы calculate_ice_integrality.

    Для сценария также можно выставить различные параметры моделирования. Можно настроить дату начала построения расписания,
     количество дней планирования, количество дней сверх плана, ограничение по времени на решение задачи и количество
     лучших маршрутов для построения расписания.
    """
    if name == 'base':
        return {'result': 'error', 'desc': f'Запрещено создавать сценарий с названием {name}'}

    if not main_data.filename.endswith('.xlsx'):
        return {'result': 'error', 'desc': f'Файл {main_data.filename} должен быть формата .xlsx'}
    if not velocity_env.filename.endswith('.xlsx'):
        return {'result': 'error', 'desc': f'Файл {velocity_env.filename} должен быть формата .xlsx'}

    scenario_dir = os.path.join('.', 'data', 'scenarios', name)
    if os.path.exists(scenario_dir):
        shutil.rmtree(scenario_dir)
    os.makedirs(scenario_dir)

    input_dir = os.path.join(scenario_dir, 'input')
    os.makedirs(input_dir)
    output_dir = os.path.join(scenario_dir, 'output')
    os.makedirs(output_dir)

    f = await main_data.read()
    xlsx = io.BytesIO(f)
    wb = openpyxl.load_workbook(xlsx)
    wb.save(os.path.join(input_dir, 'model_data.xlsx'))

    f = await velocity_env.read()
    xlsx = io.BytesIO(f)
    wb = openpyxl.load_workbook(xlsx)
    wb.save(os.path.join(input_dir, 'velocity_env.xlsx'))

    config = ScenarioConfig(
        start_date=start_date,
        duration_days=duration_days,
        interval_hours=1,
        cross_days=cross_days,
        timelimit=timelimit,
        k_bests=k_bests,
    )
    config.to_json(input_dir)

    return {'result': 'success', 'desc': f'Входные данные для расчета с названием {name} успешно загружены'}


@app.post('/create_parent_scenario/', tags=['Сценарии'])
async def create_parent_scenario(
        name: str,
        main_data: UploadFile,
        velocity_env: UploadFile,
        start_date: str = '27-02-2022',
        end_date: str = '05-05-2022',
        duration_days: int = 7,
        # interval_hours: int = 1,
        cross_days: int = 2,
        timelimit: int = 900,
        k_bests: int = 5,
):
    """
    Создание родительского сценария. Запуск такого сценария осуществляется последовательно интервалами времени,
     заданными в настройках сценария.

    Для сценария можно указать последнюю дату планирования - дату до которой будет осуществлен последовательный расчет
     расписания.
     """

    if name == 'base':
        return {'result': 'error', 'desc': f'Запрещено создавать сценарий с названием {name}'}

    if not main_data.filename.endswith('.xlsx'):
        return {'result': 'error', 'desc': f'Файл {main_data.filename} должен быть формата .xlsx'}
    if not velocity_env.filename.endswith('.xlsx'):
        return {'result': 'error', 'desc': f'Файл {velocity_env.filename} должен быть формата .xlsx'}

    scenario_dir = os.path.join('.', 'data', 'scenarios', name)
    if os.path.exists(scenario_dir):
        shutil.rmtree(scenario_dir)
    os.makedirs(scenario_dir)

    input_dir = os.path.join(scenario_dir, 'input')
    os.makedirs(input_dir)
    output_dir = os.path.join(scenario_dir, 'output')
    os.makedirs(output_dir)

    f = await main_data.read()
    xlsx = io.BytesIO(f)
    wb = openpyxl.load_workbook(xlsx)
    wb.save(os.path.join(input_dir, 'model_data.xlsx'))

    f = await velocity_env.read()
    xlsx = io.BytesIO(f)
    wb = openpyxl.load_workbook(xlsx)
    wb.save(os.path.join(input_dir, 'velocity_env.xlsx'))

    config = ParentScenarioConfig(
        start_date=start_date,
        end_date=end_date,
        duration_days=duration_days,
        interval_hours=1,
        cross_days=cross_days,
        timelimit=timelimit,
        k_bests=k_bests,
    )
    config.to_json(input_dir)

    return {'result': 'success', 'desc': f'Входные данные для расчета с названием {name} успешно загружены'}


@app.post('/download_scenario/', tags=['Сценарии'])
async def download_scenario(
        name: str = 'base',
):
    """
    Выгрузка входных и выходных (если они есть) данные сценария.
    """

    scenario_dir = os.path.join('.', 'data', 'scenarios', name)
    if not os.path.exists(scenario_dir):
        return {'result': 'error', 'desc': f'Входные данные с названием {name} не загружены'}

    tmp_zf_path = os.path.join('.', 'data', 'tmp', f'{name}.zip')
    tmp_zf_data = zipfile.ZipFile(tmp_zf_path, "w")
    for dirname, subdirs, files in os.walk(scenario_dir):
        for filename in files:
            short_dirname = dirname.replace(f'/data/scenarios/{name}', '')
            tmp_zf_data.write(os.path.join(dirname, filename), os.path.join(short_dirname, filename))
    tmp_zf_data.close()

    return_data = io.BytesIO()
    with open(tmp_zf_path, 'rb') as fo:
        return_data.write(fo.read())
    return_data.seek(0)

    os.remove(tmp_zf_path)

    return StreamingResponse(return_data, media_type="application/zip",
                             headers={'Content-Disposition': f'attachment; filename="input_{name}"'})


@app.post('/calculate_scenario/', tags=['Моделирование'])
async def calculate_scenario(
        name: str = 'base',
        is_parent: bool = False,
):
    """
    Запуск выбранного сценария для оптимизации расписания движения ледоколов по СМП.

    **is_parent** -- признак запуска в режиме родительского сценария (последовательный расчет сценариев)
    """

    scenario_dir = os.path.join('.', 'data', 'scenarios', name)
    if not os.path.exists(scenario_dir):
        return {'result': 'error', 'desc': f'Входные данные с названием {name} не загружены'}

    if is_parent:
        scenario = ParentScenario.create_scenario(scenario_folder_path=scenario_dir, scenario_name=name)
    else:
        scenario = Scenario.create_scenario(scenario_folder_path=scenario_dir, scenario_name=name)

    try:
        scenario.optimize()
    except Exception as e:
        return {'result': 'error', 'desc': f'Не удалось решить модель\n{e}'}

    return {'result': 'success', 'desc': f'Сценарий {name} успешно расчитан'}


@app.post('/calculate_ice_integrality/', tags=['Моделирование'])
async def calculate_ice_integrality(
        main_data: UploadFile,
        integer_velocity: UploadFile,
):
    """
    Запуск расчета тяжести льда. Результат запроса - таблица с интегральной тяжестью льда на каждую
    переданную дату.

    При запуске необходимо передать данные о ледовой обстановке, данные о расположении портов, а также граф переходов
    в СМП.
    """

    if not main_data.filename.endswith('.xlsx'):
        return {'result': 'error', 'desc': f'Файл {main_data.filename} должен быть формата .xlsx'}
    if not integer_velocity.filename.endswith('.xlsx'):
        return {'result': 'error', 'desc': f'Файл {integer_velocity.filename} должен быть формата .xlsx'}

    tmp_folder_name = 'iw_' + datetime.now().strftime("%Y%m%d%H")
    tmp_dir_path = os.path.join('.', 'data', 'tmp', tmp_folder_name)
    if os.path.exists(tmp_dir_path):
        shutil.rmtree(tmp_dir_path)
    os.makedirs(tmp_dir_path)

    f = await main_data.read()
    xlsx = io.BytesIO(f)
    wb = openpyxl.load_workbook(xlsx)
    wb.save(os.path.join(tmp_dir_path, 'model_data.xlsx'))

    f = await integer_velocity.read()
    xlsx = io.BytesIO(f)
    wb = openpyxl.load_workbook(xlsx)
    wb.save(os.path.join(tmp_dir_path, 'IntegrVelocity.xlsx'))

    result_df = dump_velocity_length(tmp_dir_path)
    shutil.rmtree(tmp_dir_path)
    with pd.ExcelWriter(os.path.join(tmp_dir_path, 'velocity_env.xlsx')) as writer:
        result_df.to_excel(writer)

    tmp_zf_path = os.path.join(tmp_dir_path, f'result.zip')
    tmp_zf_data = zipfile.ZipFile(tmp_zf_path, "w")
    for dirname, subdirs, files in os.walk(tmp_dir_path):
        for filename in files:
            short_dirname = dirname.replace(f'/data/tmp/{tmp_folder_name}', '')
            tmp_zf_data.write(os.path.join(dirname, filename), os.path.join(short_dirname, filename))
    tmp_zf_data.close()

    return_data = io.BytesIO()
    with open(tmp_zf_path, 'rb') as fo:
        return_data.write(fo.read())
    return_data.seek(0)

    # shutil.rmtree(tmp_dir_path)

    return StreamingResponse(return_data, media_type="application/zip",
                             headers={'Content-Disposition': f'attachment; filename="result"'})
